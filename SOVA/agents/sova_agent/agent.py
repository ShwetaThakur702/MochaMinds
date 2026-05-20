import os
import requests
import logging
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from dotenv import load_dotenv
from typing import Optional, Tuple, List

from .rules import run_r1, run_r2, run_r3, run_r4, run_r5
from .schema import SessionLocal, AppNotification

load_dotenv()
logger = logging.getLogger(__name__)

# --- Paths & Configs ---
CLUSTER_MAP_PATH = os.getenv("CLUSTER_MAP_PATH", "config/cluster_map.csv")
CURRENCY_MAP_PATH = os.getenv("CURRENCY_MAP_PATH", "config/currency_map.csv")
GEO_POLICY_PATH = os.getenv("GEO_POLICY_PATH", "config/geo_policy.csv")
JD_LIBRARY_PATH = os.getenv("JD_LIBRARY_PATH", "data/JDs_Consolidated_1_-_Copy.xlsx")
TEAMS_WEBHOOK_URL = os.getenv("TEAMS_WEBHOOK_URL", "")
GOOGLE_CREDS_PATH = os.getenv("GOOGLE_CREDS_PATH", "config/google_credentials.json")
GOOGLE_SHARE_EMAIL = os.getenv("GOOGLE_SHARE_EMAIL", "")
GOOGLE_SPREADSHEET_ID = os.getenv("GOOGLE_SPREADSHEET_ID", "")

DATE_COLS_SO = ["SO Creation Date", "SO Submission Date", "Start Date"]
REQUIRED_SO_COLS = [
    "SO ID", "Status", "SO Creation Date", "SO Submission Date",
    "Start Date", "Hiring Geo/Location", "Requirement Location",
    "Country", "City Name", "SL/IND_CLUSTER", "PSID of Hiring Manager",
    "Customer Name", "Project ID", "Description", "Project Role",
    "Primary Skill Description", "Keywords", "Technology"
]
SEVERITY_COLORS = {
    "High": "FFCCCC",
    "Medium": "FFF2CC",
    "Low": "E2EFDA",
}

OUTPUT_COLUMNS = [
    "SO ID", "Project Code", "Customer Name", "Service Order Status",
    "Rule ID", "Rule Name", "Exception Reason", "Severity",
    "PSID of Hiring Manager",
    "Expected Billing Start Date", "SO Creation Date", "SO Submission Date",
    "Hiring GeoLocation", "Onsite/Offshore", "Work Location Country", "Work Location City",
    "SO ClusterGroup", "Manager ClusterGroup",
    "Budgeted CTC Currency", "Recommended Currency",
    "Job Title", "Primary Skill Set", "Description Quality Flag",
    "Recommended Action", "Validity Indicator", "Run Date",
]

# ==========================================
# CLEANER FUNCTIONS
# ==========================================
def clean_so(df: pd.DataFrame) -> Tuple[pd.DataFrame, list, list]:
    warnings = []
    rejected = []
    missing_cols = [c for c in REQUIRED_SO_COLS if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing mandatory SO columns: {missing_cols}")

    str_cols = df.select_dtypes(include="object").columns
    df[str_cols] = df[str_cols].apply(lambda col: col.str.strip() if col.dtype == "object" else col)

    null_so = df[df["SO ID"].isnull()]
    if not null_so.empty:
        rejected.extend(null_so.index.tolist())
        warnings.append(f"{len(null_so)} rows rejected: null SO ID")
        df = df[df["SO ID"].notna()].copy()

    df["SO ID"] = df["SO ID"].astype(str).str.strip()

    for col in DATE_COLS_SO:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=False)
            bad = df[df[col].isnull() & df["SO ID"].notna()]
            if not bad.empty:
                warnings.append(f"{len(bad)} rows have unparseable '{col}' — excluded from R1")

    df["Status"] = df["Status"].str.strip()
    df["Requirement Location"] = df["Requirement Location"].str.strip().str.capitalize()
    df["Country"] = df["Country"].str.strip()
    df["Hiring Geo/Location"] = df["Hiring Geo/Location"].str.strip().str.upper()
    df["SL/IND_CLUSTER"] = df["SL/IND_CLUSTER"].str.strip()

    active_statuses = ["Active", "OnHold-TA"]
    inactive = df[~df["Status"].isin(active_statuses)]
    if not inactive.empty:
        warnings.append(f"{len(inactive)} rows skipped: status not in {active_statuses}")
        df = df[df["Status"].isin(active_statuses)].copy()

    for col in ["Description", "Keywords", "Primary Skill Description", "Project Role"]:
        if col in df.columns:
            df[col] = df[col].fillna("N/A")

    logger.info(f"SO cleaning done: {len(df)} valid rows, {len(rejected)} rejected")
    return df, rejected, warnings

def clean_ris(df: pd.DataFrame) -> Tuple[pd.DataFrame, list]:
    warnings = []
    str_cols = df.select_dtypes(include="object").columns
    df[str_cols] = df[str_cols].apply(lambda col: col.str.strip() if col.dtype == "object" else col)
    if "Project Mgr PSID" in df.columns:
        df["Project Mgr PSID"] = pd.to_numeric(df["Project Mgr PSID"], errors="coerce")
        null_psid = df["Project Mgr PSID"].isnull().sum()
        if null_psid:
            warnings.append(f"RIS: {null_psid} rows with null Project Mgr PSID")
    if "SL/IND_CLUSTER" in df.columns:
        df["SL/IND_CLUSTER"] = df["SL/IND_CLUSTER"].str.strip()
    logger.info(f"RIS cleaning done: {len(df)} rows")
    return df, warnings


# ==========================================
# REPORTER FUNCTIONS
# ==========================================
def generate_report(exceptions, run_date, output_dir="/tmp"):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = "SOVA_ValidityReport_{}.xlsx".format(timestamp)
    filepath = os.path.join(output_dir, filename)

    if not exceptions:
        df = pd.DataFrame(columns=OUTPUT_COLUMNS)
    else:
        df = pd.DataFrame(exceptions)
        for col in OUTPUT_COLUMNS:
            if col not in df.columns:
                df[col] = "N/A"
        df = df[OUTPUT_COLUMNS]

    df = df.fillna("N/A")
    df = df.replace("", "N/A")
    df = df.replace("nan", "N/A")
    df = df.replace("NaT", "N/A")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOVA_Exceptions"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row_idx, row in df.iterrows():
        severity = str(row.get("Severity", ""))
        row_fill_color = SEVERITY_COLORS.get(severity, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=row_fill_color)
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            val = row[col_name]
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=str(val))
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            cell.fill = row_fill

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(col_name), 15)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    summary_data = [
        ["SOVA Validity Report — Summary", ""],
        ["Run Date", run_date],
        ["Total Exceptions", len(df)],
        ["", ""],
        ["Rule", "Exception Count"],
    ]
    rule_counts = df["Rule ID"].value_counts().to_dict() if not df.empty else {}
    for r in ["R1", "R2", "R3", "R4", "R5"]:
        summary_data.append([r, rule_counts.get(r, 0)])

    for row in summary_data:
        ws2.append(row)

    ws2["A1"].font = Font(name="Arial", bold=True, size=12)
    wb.save(filepath)
    return filepath


# ==========================================
# NOTIFIER FUNCTIONS
# ==========================================
def send_ops_alert(run_id: int, summary: dict):
    total = summary.get("total_exceptions", 0)
    if total == 0:
        message = f"✅ Run #{run_id} completed successfully. 0 exceptions found."
        severity = "INFO"
    else:
        message = f"🚨 Run #{run_id} completed with {total} exceptions! R1: {summary.get('r1_count',0)}, R2: {summary.get('r2_count',0)}. Review required."
        severity = "HIGH" if total > 50 else "WARNING"

    try:
        db = SessionLocal()
        notif = AppNotification(run_id=run_id, message=message, severity=severity)
        db.add(notif)
        db.commit()
        db.close()
    except Exception as e:
        logger.error(f"Failed to save AppNotification: {e}")

    if not TEAMS_WEBHOOK_URL:
        logger.info("TEAMS_WEBHOOK_URL not set. Skipping external webhook alert.")
        return

    payload = {
        "text": "SOVA Alert",
        "attachments": [{
            "title": f"SOVA Validation Run #{run_id}",
            "text": message,
            "color": "#FF0000" if severity == "HIGH" else "#FFA500" if severity == "WARNING" else "#00FF00",
            "fields": [
                {"title": "Total SOs", "value": summary.get("total_sos", 0), "short": True},
                {"title": "Exceptions", "value": total, "short": True}
            ]
        }]
    }
    try:
        resp = requests.post(TEAMS_WEBHOOK_URL, json=payload, timeout=3)
        logger.info(f"Webhook alert sent to {TEAMS_WEBHOOK_URL} - Status {resp.status_code}")
    except Exception as e:
        logger.error(f"Failed to send webhook alert: {e}")


# ==========================================
# GOOGLE SHEETS FUNCTIONS
# ==========================================
def _get_client():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise ImportError("gspread and google-auth are not installed. Run: pip install gspread google-auth")

    if not os.path.exists(GOOGLE_CREDS_PATH):
        raise FileNotFoundError(f"Google credentials file not found at '{GOOGLE_CREDS_PATH}'.")

    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_file(GOOGLE_CREDS_PATH, scopes=scopes)
    return gspread.authorize(creds)

def _write_run_to_spreadsheet(spreadsheet, exceptions: list, summary: dict, run_id: Optional[int], is_new_spreadsheet: bool = False) -> str:
    run_label  = f"Run_{run_id}" if run_id else "Run"
    ts         = datetime.now().strftime("%d%b%Y_%I-%M%p")
    tab_name   = f"{run_label}_{ts}"
    total_rows = 20 + max(len(exceptions), 1) + 5
    
    if is_new_spreadsheet:
        ws = spreadsheet.sheet1
        ws.update_title(tab_name)
    else:
        ws = spreadsheet.add_worksheet(title=tab_name, rows=total_rows, cols=30)

    summary_rows = [
        ["SOVA Validity Report — Summary", "", "", "", ""],
        ["Run Date",         summary.get("run_date",         "N/A")],
        ["SO File",          summary.get("so_file",          "N/A")],
        ["RIS File",         summary.get("ris_file",         "N/A")],
        ["Total SOs",        summary.get("total_sos",        "N/A")],
        ["Rejected Rows",    summary.get("rejected_rows",    "N/A")],
        ["Total Exceptions", summary.get("total_exceptions", "N/A")],
        [],
        ["Rule", "Exceptions"],
        ["R1 — Billing Start Date",  summary.get("r1_count", 0)],
        ["R2 — Geo / Location",      summary.get("r2_count", 0)],
        ["R3 — Cluster Group",       summary.get("r3_count", 0)],
        ["R4 — Currency Mapping",    summary.get("r4_count", 0)],
        ["R5 — JD Quality",          summary.get("r5_count", 0)],
        [],
        ["Skipped Rules", ", ".join(summary.get("skipped_rules", [])) or "None"],
        [], []
    ]

    if exceptions:
        headers = list(exceptions[0].keys())
        exc_rows = [headers] + [[str(row.get(h, "N/A")) for h in headers] for row in exceptions]
    else:
        headers = ["SO ID", "Rule ID", "Rule Name", "Exception Reason", "Severity", "Recommended Action", "Validity Indicator", "Run Date"]
        exc_rows = [headers, ["No exceptions found for this run."]]

    all_rows = summary_rows + exc_rows
    ws.update(all_rows, "A1")

    ws.format("A1:B1", {"textFormat": {"bold": True, "fontSize": 12}})
    ws.format("A9:B9", {"textFormat": {"bold": True}})
    
    exc_header_row = len(summary_rows) + 1
    ws.format(f"A{exc_header_row}:Z{exc_header_row}", {
        "textFormat": {"bold": True},
        "backgroundColor": {"red": 0.9, "green": 0.9, "blue": 0.9}
    })
    ws.freeze(rows=exc_header_row)

    logger.info(f"Wrote combined tab '{tab_name}' to spreadsheet {spreadsheet.url}")
    return spreadsheet.url

def export_to_sheets(exceptions: list, summary: dict, run_id: Optional[int] = None) -> str:
    client = _get_client()
    if GOOGLE_SPREADSHEET_ID:
        logger.info(f"Opening existing spreadsheet ID: {GOOGLE_SPREADSHEET_ID}")
        try:
            spreadsheet = client.open_by_key(GOOGLE_SPREADSHEET_ID)
        except Exception as e:
            raise Exception(f"Could not open spreadsheet. Error: {e}")
        url = _write_run_to_spreadsheet(spreadsheet, exceptions, summary, run_id)
    else:
        run_label   = f"Run #{run_id}" if run_id else "Run"
        sheet_title = f"SOVA Validity Report — {run_label} — {summary.get('run_date', '')}"[:100]
        spreadsheet = client.create(sheet_title)
        logger.info(f"Google Sheet created: {spreadsheet.url}")
        url = _write_run_to_spreadsheet(spreadsheet, exceptions, summary, run_id, is_new_spreadsheet=True)
        if GOOGLE_SHARE_EMAIL:
            spreadsheet.share(GOOGLE_SHARE_EMAIL, perm_type="user", role="writer")
        spreadsheet.share("", perm_type="anyone", role="reader")
        url = spreadsheet.url
    return url


# ==========================================
# AGENT RUNNER
# ==========================================
def _load_ref(path: str, name: str) -> Optional[pd.DataFrame]:
    if not os.path.exists(path):
        logger.warning(f"{name} not found at '{path}' — dependent rule will be skipped")
        return None
    try:
        df = pd.read_csv(path)
        logger.info(f"{name} loaded: {len(df)} rows")
        return df
    except Exception as e:
        logger.warning(f"Failed to load {name}: {e} — dependent rule will be skipped")
        return None

def run_agent(so_bytes: bytes, ris_bytes: bytes, so_filename: str, ris_filename: str, output_dir: str = "reports") -> dict:
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logs = []
    skipped_rules = []
    all_exceptions = []

    def log(msg, level="INFO"):
        logger.info(msg)
        logs.append({"level": level, "message": msg})

    try:
        so_raw = pd.read_excel(so_bytes, dtype=str)
        log(f"SO file loaded: {len(so_raw)} rows")
    except Exception as e:
        raise ValueError(f"Failed to read SO file: {e}")

    try:
        ris_raw = pd.read_excel(ris_bytes, dtype=str)
        log(f"RIS file loaded: {len(ris_raw)} rows")
    except Exception as e:
        raise ValueError(f"Failed to read RIS file: {e}")

    so_df, rejected, so_warnings = clean_so(so_raw)
    for w in so_warnings:
        log(w, "WARNING")

    ris_df, ris_warnings = clean_ris(ris_raw)
    for w in ris_warnings:
        log(w, "WARNING")

    total_sos = len(so_df)
    log(f"After cleaning: {total_sos} valid SO rows, {len(rejected)} rejected")

    geo_policy = _load_ref(GEO_POLICY_PATH, "Geo Policy Map")
    cluster_map = _load_ref(CLUSTER_MAP_PATH, "Cluster Map")
    currency_map = _load_ref(CURRENCY_MAP_PATH, "Currency Map")

    jd_library = None
    if os.path.exists(JD_LIBRARY_PATH):
        try:
            jd_library = pd.read_excel(JD_LIBRARY_PATH, dtype=str)
            log(f"JD Library loaded: {len(jd_library)} rows from '{JD_LIBRARY_PATH}'")
        except Exception as e:
            log(f"JD Library could not be loaded (R5 will use heuristic-only mode): {e}", "WARNING")
    else:
        log(f"JD Library not found at '{JD_LIBRARY_PATH}' — R5 will use heuristic-only mode.", "WARNING")

    r1 = run_r1(so_df, run_date)
    all_exceptions.extend(r1)
    log(f"R1 (Billing Start Date): {len(r1)} exceptions")

    r2, r2_skipped = run_r2(so_df, geo_policy, run_date)
    all_exceptions.extend(r2)
    if r2_skipped:
        skipped_rules.append("R2")
        log("R2 skipped: geo_policy.csv missing", "WARNING")
    else:
        log(f"R2 (Geo/Location): {len(r2)} exceptions")

    r3, r3_skipped = run_r3(so_df, cluster_map, run_date)
    all_exceptions.extend(r3)
    if r3_skipped:
        skipped_rules.append("R3")
        log("R3 skipped: cluster_map.csv missing", "WARNING")
    else:
        log(f"R3 (Cluster Group): {len(r3)} exceptions")

    r4, r4_skipped = run_r4(so_df, currency_map, run_date)
    all_exceptions.extend(r4)
    if r4_skipped:
        skipped_rules.append("R4")
        log("R4 skipped: currency_map.csv missing", "WARNING")
    else:
        log(f"R4 (Currency): {len(r4)} exceptions")

    r5 = run_r5(so_df, run_date, jd_library=jd_library)
    all_exceptions.extend(r5)
    log(f"R5 (JD Quality): {len(r5)} exceptions")

    output_path = generate_report(all_exceptions, run_date, output_dir)
    log(f"Report generated: {output_path}")

    return {
        "run_date": run_date,
        "so_file": so_filename,
        "ris_file": ris_filename,
        "total_sos": total_sos,
        "rejected_rows": len(rejected),
        "total_exceptions": len(all_exceptions),
        "r1_count": len(r1),
        "r2_count": len(r2),
        "r3_count": len(r3),
        "r4_count": len(r4),
        "r5_count": len(r5),
        "skipped_rules": skipped_rules,
        "output_path": output_path,
        "logs": logs,
        "exceptions": all_exceptions,
    }
